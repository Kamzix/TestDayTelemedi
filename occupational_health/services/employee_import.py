from dataclasses import dataclass, field
from datetime import datetime
from zipfile import BadZipFile

from django.core.exceptions import ValidationError
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import load_workbook

from occupational_health.models import Employee


EXPECTED_HEADERS = [
    'first_name',
    'last_name',
    'pesel',
    'birth_date',
    'identity_document',
    'email',
    'phone',
    'city',
    'street',
    'building_number',
    'apartment_number',
    'job_position',
    'active',
]


@dataclass
class ImportErrorItem:
    row: int
    error: str


@dataclass
class ImportResult:
    created_count: int = 0
    errors: list[ImportErrorItem] = field(default_factory=list)


def import_employees_from_xlsx(file_obj, organization):
    result = ImportResult()
    try:
        workbook = load_workbook(file_obj, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError):
        result.errors.append(ImportErrorItem(
            row=0,
            error='Nie mozna odczytac pliku XLSX. Sprawdz format pliku.',
        ))
        return result
    sheet = workbook.active

    headers = [_normalize_header(cell.value) for cell in sheet[1]]
    missing_headers = [header for header in EXPECTED_HEADERS if header not in headers]
    if missing_headers:
        result.errors.append(ImportErrorItem(
            row=1,
            error=f'Brak wymaganych kolumn: {", ".join(missing_headers)}.',
        ))
        return result

    header_indexes = {header: headers.index(header) for header in EXPECTED_HEADERS}
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if _is_empty_row(row):
            continue

        data = {
            header: _clean_value(row[header_indexes[header]])
            for header in EXPECTED_HEADERS
        }
        try:
            employee = Employee(
                organization=organization,
                first_name=data['first_name'],
                last_name=data['last_name'],
                pesel=data['pesel'],
                birth_date=_parse_date(data['birth_date']),
                identity_document=data['identity_document'],
                email=data['email'],
                phone=data['phone'],
                city=data['city'],
                street=data['street'],
                building_number=data['building_number'],
                apartment_number=data['apartment_number'],
                job_position=data['job_position'],
                active=_parse_bool(data['active']),
            )
            employee.full_clean()
            employee.save()
            result.created_count += 1
        except (TypeError, ValueError, ValidationError) as error:
            result.errors.append(ImportErrorItem(
                row=row_number,
                error=_format_error(error),
            ))

    return result


def _normalize_header(value):
    return str(value).strip() if value is not None else ''


def _clean_value(value):
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    return value


def _is_empty_row(row):
    return all(value is None or str(value).strip() == '' for value in row)


def _parse_date(value):
    if value in (None, ''):
        return None
    if hasattr(value, 'date'):
        return value.date()
    if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
        return value
    if isinstance(value, str):
        return datetime.strptime(value, '%Y-%m-%d').date()
    raise ValueError('Nieprawidlowa data. Uzyj formatu YYYY-MM-DD.')


def _parse_bool(value):
    if value in (None, ''):
        return True
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in {'true', '1', 'yes', 'tak'}:
        return True
    if normalized in {'false', '0', 'no', 'nie'}:
        return False
    raise ValueError('Pole active musi miec wartosc true/false, 1/0, yes/no albo tak/nie.')


def _format_error(error):
    if isinstance(error, ValidationError):
        if hasattr(error, 'message_dict'):
            messages = []
            for field, field_errors in error.message_dict.items():
                messages.append(f'{field}: {", ".join(field_errors)}')
            return '; '.join(messages)
        return '; '.join(error.messages)
    return str(error)
